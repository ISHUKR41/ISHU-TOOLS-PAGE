"""
ISHU TOOLS — Office / Document Tools (2026 batch #3)
Adds: csv-to-excel, excel-to-csv, pdf-page-extractor
"""
from __future__ import annotations

import csv
import io
import re
from pathlib import Path
from typing import Any

from .handlers import ExecutionResult


def _safe(name: str, limit: int = 60) -> str:
    cleaned = re.sub(r"[^\w\s-]", "", name or "file")[:limit].strip()
    return cleaned.replace(" ", "_") or "file"


# ─── 1. CSV to Excel ─────────────────────────────────────────────────────────

def _handle_csv_to_excel(files: list[Path], payload: dict[str, Any], job_dir: Path) -> ExecutionResult:
    if not files:
        return ExecutionResult(kind="json", message="Please upload a CSV file.", data={"error": "No file"})

    src = files[0]
    delimiter = payload.get("delimiter") or ","
    if delimiter == "tab":
        delimiter = "\t"
    elif delimiter == "semicolon":
        delimiter = ";"
    elif delimiter == "pipe":
        delimiter = "|"

    try:
        from openpyxl import Workbook
    except ImportError:
        return ExecutionResult(kind="json", message="Excel engine not installed on server.", data={"error": "openpyxl missing"})

    try:
        text = src.read_text(encoding="utf-8", errors="replace")
    except Exception:
        try:
            text = src.read_text(encoding="latin-1", errors="replace")
        except Exception as e:
            return ExecutionResult(kind="json", message=f"Could not read CSV: {str(e)[:200]}", data={"error": str(e)[:500]})

    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    if not rows:
        return ExecutionResult(kind="json", message="CSV file is empty.", data={"error": "empty"})

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for row in rows:
        ws.append(row)

    # Auto-size columns (best effort, capped to 60 chars)
    for col in ws.columns:
        try:
            max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(60, max(8, max_len + 2))
        except Exception:
            pass

    out_path = job_dir / f"{src.stem}.xlsx"
    wb.save(str(out_path))

    return ExecutionResult(
        kind="file",
        message=f"✅ Converted {len(rows)} rows to Excel",
        output_path=out_path,
        filename=f"{_safe(src.stem)}.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ─── 2. Excel to CSV ─────────────────────────────────────────────────────────

def _handle_excel_to_csv(files: list[Path], payload: dict[str, Any], job_dir: Path) -> ExecutionResult:
    if not files:
        return ExecutionResult(kind="json", message="Please upload an Excel (.xlsx) file.", data={"error": "No file"})

    src = files[0]
    sheet_name = (payload.get("sheet") or "").strip()

    try:
        from openpyxl import load_workbook
    except ImportError:
        return ExecutionResult(kind="json", message="Excel engine not installed on server.", data={"error": "openpyxl missing"})

    try:
        wb = load_workbook(filename=str(src), data_only=True, read_only=True)
    except Exception as e:
        return ExecutionResult(kind="json", message=f"Could not open Excel file: {str(e)[:200]}", data={"error": str(e)[:500]})

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        used_sheet = sheet_name
    else:
        ws = wb.active
        used_sheet = ws.title

    out_path = job_dir / f"{src.stem}.csv"
    row_count = 0
    with out_path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in row])
            row_count += 1

    return ExecutionResult(
        kind="file",
        message=f"✅ Extracted sheet '{used_sheet}' — {row_count} rows",
        output_path=out_path,
        filename=f"{_safe(src.stem)}.csv",
        content_type="text/csv",
    )


# ─── 3. PDF Page Extractor ───────────────────────────────────────────────────

def _parse_pages(spec: str, total: int) -> list[int]:
    """Parse '1,3-5,8' into [1,3,4,5,8] (1-indexed). Returns sorted unique list within range."""
    out: set[int] = set()
    spec = (spec or "").strip()
    if not spec:
        return []
    for part in spec.split(","):
        part = part.strip()
        if not part:
            continue
        if "-" in part:
            try:
                a, b = part.split("-", 1)
                a_i = int(a.strip())
                b_i = int(b.strip())
                if a_i > b_i:
                    a_i, b_i = b_i, a_i
                for n in range(a_i, b_i + 1):
                    if 1 <= n <= total:
                        out.add(n)
            except ValueError:
                continue
        else:
            try:
                n = int(part)
                if 1 <= n <= total:
                    out.add(n)
            except ValueError:
                continue
    return sorted(out)


def _handle_pdf_page_extractor(files: list[Path], payload: dict[str, Any], job_dir: Path) -> ExecutionResult:
    if not files:
        return ExecutionResult(kind="json", message="Please upload a PDF file.", data={"error": "No file"})

    src = files[0]
    pages_spec = (payload.get("pages") or "").strip()
    if not pages_spec:
        return ExecutionResult(
            kind="json",
            message="Please enter pages to extract (e.g. '1,3-5,8' for pages 1, 3, 4, 5 and 8).",
            data={"error": "No pages"},
        )

    try:
        from pypdf import PdfReader, PdfWriter
    except ImportError:
        return ExecutionResult(kind="json", message="PDF engine not installed on server.", data={"error": "pypdf missing"})

    try:
        reader = PdfReader(str(src))
    except Exception as e:
        return ExecutionResult(kind="json", message=f"Could not open PDF: {str(e)[:200]}", data={"error": str(e)[:500]})

    total = len(reader.pages)
    if total == 0:
        return ExecutionResult(kind="json", message="PDF has no pages.", data={"error": "empty PDF"})

    pages = _parse_pages(pages_spec, total)
    if not pages:
        return ExecutionResult(
            kind="json",
            message=f"No valid pages found in your input. PDF has {total} pages — use numbers between 1 and {total}.",
            data={"error": "Invalid pages", "total_pages": total},
        )

    writer = PdfWriter()
    for n in pages:
        writer.add_page(reader.pages[n - 1])

    out_path = job_dir / f"{src.stem}_extracted.pdf"
    with out_path.open("wb") as f:
        writer.write(f)

    pages_str = ", ".join(str(p) for p in pages[:10])
    if len(pages) > 10:
        pages_str += f" (+{len(pages) - 10} more)"

    return ExecutionResult(
        kind="file",
        message=f"✅ Extracted {len(pages)} pages — {pages_str}",
        output_path=out_path,
        filename=f"{_safe(src.stem)}_pages.pdf",
        content_type="application/pdf",
    )


# ─── Registry ────────────────────────────────────────────────────────────────

OFFICE_HANDLERS = {
    "csv-to-excel": _handle_csv_to_excel,
    "csv-to-xlsx": _handle_csv_to_excel,
    "excel-to-csv": _handle_excel_to_csv,
    "xlsx-to-csv": _handle_excel_to_csv,
    "pdf-page-extractor": _handle_pdf_page_extractor,
    "extract-pdf-pages": _handle_pdf_page_extractor,
}
